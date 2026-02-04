from __future__ import annotations

from dataclasses import dataclass
from io import BytesIO
import sqlite3
from typing import Iterable, List, Optional, Tuple

from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter


@dataclass(frozen=True)
class RevenueSummary:
    total: int
    count: int
    cash: int
    transfer: int
    qr: int


@dataclass(frozen=True)
class ExpenseSummary:
    total: int
    cash: int
    transfer: int
    qr: int
    categories: List[Tuple[str, int]]
    other_amount: int


@dataclass(frozen=True)
class AttendanceSummary:
    booked: int
    attended: int
    noshow: int
    cancelled: int


def _int_or_zero(value: Optional[int]) -> int:
    return int(value or 0)


def get_revenue_summary(db_path: str, date_from: str, date_to: str) -> RevenueSummary:
    with sqlite3.connect(db_path) as conn:
        cur = conn.execute(
            """
            SELECT
              COALESCE(SUM(amount), 0) AS total_amount,
              COUNT(*) AS total_count
            FROM payments
            WHERE status = 'paid'
              AND method IN ('cash','transfer','qr')
              AND pay_date BETWEEN ? AND ?
            """,
            (date_from, date_to),
        )
        row = cur.fetchone() or (0, 0)
        total_amount = _int_or_zero(row[0])
        total_count = _int_or_zero(row[1])

        cur = conn.execute(
            """
            SELECT method, COALESCE(SUM(amount), 0) AS total_amount
            FROM payments
            WHERE status = 'paid'
              AND method IN ('cash','transfer','qr')
              AND pay_date BETWEEN ? AND ?
            GROUP BY method
            """,
            (date_from, date_to),
        )
        by_method = {row[0]: _int_or_zero(row[1]) for row in cur.fetchall()}

    return RevenueSummary(
        total=total_amount,
        count=total_count,
        cash=by_method.get("cash", 0),
        transfer=by_method.get("transfer", 0),
        qr=by_method.get("qr", 0),
    )


def list_paid_payments(
    db_path: str, date_from: str, date_to: str
) -> List[Tuple[str, str, str, str, int, str]]:
    with sqlite3.connect(db_path) as conn:
        cur = conn.execute(
            """
            SELECT
              p.pay_date,
              COALESCE(c.full_name, '—') AS client_name,
              COALESCE(g.name, '—') AS group_name,
              p.purpose,
              p.amount,
              p.method
            FROM payments p
            LEFT JOIN clients c ON c.client_id = p.client_id
            LEFT JOIN groups g ON g.group_id = p.group_id
            WHERE p.status = 'paid'
              AND p.method IN ('cash','transfer','qr')
              AND p.pay_date BETWEEN ? AND ?
            ORDER BY p.pay_date DESC, p.pay_id DESC
            """,
            (date_from, date_to),
        )
        return cur.fetchall()


def get_expense_summary(db_path: str, date_from: str, date_to: str) -> ExpenseSummary:
    with sqlite3.connect(db_path) as conn:
        cur = conn.execute(
            """
            SELECT
              COALESCE(SUM(amount), 0) AS total_amount
            FROM expenses
            WHERE exp_date BETWEEN ? AND ?
            """,
            (date_from, date_to),
        )
        total_amount = _int_or_zero(cur.fetchone()[0])

        cur = conn.execute(
            """
            SELECT method, COALESCE(SUM(amount), 0) AS total_amount
            FROM expenses
            WHERE exp_date BETWEEN ? AND ?
            GROUP BY method
            """,
            (date_from, date_to),
        )
        by_method = {row[0]: _int_or_zero(row[1]) for row in cur.fetchall()}

        cur = conn.execute(
            """
            SELECT c.name, COALESCE(SUM(e.amount), 0) AS total_amount
            FROM expenses e
            JOIN expense_categories c ON c.category_id = e.category_id
            WHERE e.exp_date BETWEEN ? AND ?
            GROUP BY c.name
            ORDER BY total_amount DESC
            """,
            (date_from, date_to),
        )
        categories = [(row[0], _int_or_zero(row[1])) for row in cur.fetchall()]

    top = categories[:5]
    other_amount = sum(amount for _, amount in categories[5:])
    return ExpenseSummary(
        total=total_amount,
        cash=by_method.get("cash", 0),
        transfer=by_method.get("transfer", 0),
        qr=by_method.get("qr", 0),
        categories=top,
        other_amount=other_amount,
    )


def list_expenses_for_period(
    db_path: str, date_from: str, date_to: str
) -> List[Tuple[str, str, int, str, Optional[str]]]:
    with sqlite3.connect(db_path) as conn:
        cur = conn.execute(
            """
            SELECT
              e.exp_date,
              c.name,
              e.amount,
              e.method,
              e.comment
            FROM expenses e
            JOIN expense_categories c ON c.category_id = e.category_id
            WHERE e.exp_date BETWEEN ? AND ?
            ORDER BY e.exp_date DESC, e.expense_id DESC
            """,
            (date_from, date_to),
        )
        return cur.fetchall()


def get_attendance_summary(db_path: str, date_from: str, date_to: str) -> AttendanceSummary:
    with sqlite3.connect(db_path) as conn:
        cur = conn.execute(
            """
            SELECT status, COUNT(*) AS total_count
            FROM visits
            WHERE visit_date BETWEEN ? AND ?
            GROUP BY status
            """,
            (date_from, date_to),
        )
        rows = {row[0]: _int_or_zero(row[1]) for row in cur.fetchall()}
    return AttendanceSummary(
        booked=rows.get("booked", 0),
        attended=rows.get("attended", 0),
        noshow=rows.get("noshow", 0),
        cancelled=rows.get("cancelled", 0),
    )


def list_attended_today_by_group(
    db_path: str, group_id: int, visit_date: str
) -> List[Tuple[str, str]]:
    with sqlite3.connect(db_path) as conn:
        cur = conn.execute(
            """
            SELECT c.full_name, c.phone
            FROM visits v
            JOIN clients c ON c.client_id = v.client_id
            WHERE v.group_id = ?
              AND v.visit_date = ?
              AND v.status = 'attended'
            ORDER BY c.full_name COLLATE NOCASE
            """,
            (group_id, visit_date),
        )
        return cur.fetchall()


def list_active_passes_today(
    db_path: str, today: str
) -> List[Tuple[str, str, str, str]]:
    with sqlite3.connect(db_path) as conn:
        cur = conn.execute(
            """
            SELECT c.full_name, g.name, p.start_date, p.end_date
            FROM passes p
            JOIN clients c ON c.client_id = p.client_id
            JOIN groups g ON g.group_id = p.group_id
            WHERE p.is_active = 1
              AND p.start_date <= ?
              AND p.end_date >= ?
            ORDER BY p.end_date ASC, c.full_name COLLATE NOCASE
            """,
            (today, today),
        )
        return cur.fetchall()


def list_passes_expiring(
    db_path: str, date_from: str, date_to: str
) -> List[Tuple[str, str, str]]:
    with sqlite3.connect(db_path) as conn:
        cur = conn.execute(
            """
            SELECT c.full_name, g.name, p.end_date
            FROM passes p
            JOIN clients c ON c.client_id = p.client_id
            JOIN groups g ON g.group_id = p.group_id
            WHERE p.is_active = 1
              AND p.end_date BETWEEN ? AND ?
            ORDER BY p.end_date ASC, c.full_name COLLATE NOCASE
            """,
            (date_from, date_to),
        )
        return cur.fetchall()


def list_clients_without_active_pass(
    db_path: str, today: str
) -> List[Tuple[str, str]]:
    with sqlite3.connect(db_path) as conn:
        cur = conn.execute(
            """
            SELECT c.full_name, g.name
            FROM client_groups cg
            JOIN clients c ON c.client_id = cg.client_id
            JOIN groups g ON g.group_id = cg.group_id
            WHERE cg.status = 'active'
              AND g.is_active = 1
              AND (cg.until_date IS NULL OR cg.until_date >= ?)
              AND NOT EXISTS (
                SELECT 1
                FROM passes p
                WHERE p.client_id = cg.client_id
                  AND p.group_id = cg.group_id
                  AND p.is_active = 1
                  AND p.start_date <= ?
                  AND p.end_date >= ?
              )
            ORDER BY c.full_name COLLATE NOCASE
            """,
            (today, today, today),
        )
        return cur.fetchall()


def count_single_visits(
    db_path: str, date_from: str, date_to: str
) -> int:
    with sqlite3.connect(db_path) as conn:
        cur = conn.execute(
            """
            SELECT COUNT(*)
            FROM visits v
            WHERE v.visit_date BETWEEN ? AND ?
              AND v.status IN ('booked','attended')
              AND NOT EXISTS (
                SELECT 1
                FROM passes p
                WHERE p.client_id = v.client_id
                  AND p.group_id = v.group_id
                  AND p.is_active = 1
                  AND p.start_date <= v.visit_date
                  AND p.end_date >= v.visit_date
              )
            """,
            (date_from, date_to),
        )
        return _int_or_zero(cur.fetchone()[0])


def list_single_visits_without_payment(
    db_path: str, date_from: str, date_to: str, limit: Optional[int] = None
) -> List[Tuple[str, str, str, str]]:
    limit_clause = ""
    params: list = [date_from, date_to]
    if limit is not None:
        limit_clause = "LIMIT ?"
        params.append(limit)
    with sqlite3.connect(db_path) as conn:
        cur = conn.execute(
            f"""
            SELECT v.visit_date, c.full_name, g.name, v.status
            FROM visits v
            JOIN clients c ON c.client_id = v.client_id
            JOIN groups g ON g.group_id = v.group_id
            WHERE v.visit_date BETWEEN ? AND ?
              AND v.status IN ('booked','attended')
              AND NOT EXISTS (
                SELECT 1
                FROM passes p
                WHERE p.client_id = v.client_id
                  AND p.group_id = v.group_id
                  AND p.is_active = 1
                  AND p.start_date <= v.visit_date
                  AND p.end_date >= v.visit_date
              )
              AND NOT EXISTS (
                SELECT 1
                FROM payments pay
                WHERE pay.visit_id = v.visit_id
                  AND pay.purpose = 'single'
                  AND pay.status != 'cancelled'
              )
            ORDER BY v.visit_date DESC, v.visit_id DESC
            {limit_clause}
            """,
            tuple(params),
        )
        return cur.fetchall()


def count_single_visits_without_payment(db_path: str, date_from: str, date_to: str) -> int:
    with sqlite3.connect(db_path) as conn:
        cur = conn.execute(
            """
            SELECT COUNT(*)
            FROM visits v
            WHERE v.visit_date BETWEEN ? AND ?
              AND v.status IN ('booked','attended')
              AND NOT EXISTS (
                SELECT 1
                FROM passes p
                WHERE p.client_id = v.client_id
                  AND p.group_id = v.group_id
                  AND p.is_active = 1
                  AND p.start_date <= v.visit_date
                  AND p.end_date >= v.visit_date
              )
              AND NOT EXISTS (
                SELECT 1
                FROM payments pay
                WHERE pay.visit_id = v.visit_id
                  AND pay.purpose = 'single'
                  AND pay.status != 'cancelled'
              )
            """,
            (date_from, date_to),
        )
        return _int_or_zero(cur.fetchone()[0])


def get_deferred_summary(
    db_path: str,
    date_from: str,
    date_to: str,
    today: str,
    overdue_days: int,
) -> Tuple[int, int, List[Tuple[int, str, str, int, str, Optional[str]]], List[Tuple[int, str, str, int, str]]]:
    with sqlite3.connect(db_path) as conn:
        cur = conn.execute(
            """
            SELECT COUNT(*), COALESCE(SUM(amount), 0)
            FROM payments
            WHERE status = 'deferred'
              AND method = 'defer'
              AND date(created_at) BETWEEN ? AND ?
            """,
            (date_from, date_to),
        )
        total_count, total_amount = cur.fetchone() or (0, 0)

        cur = conn.execute(
            """
            SELECT p.pay_id, COALESCE(c.full_name, '—') AS client_name,
                   COALESCE(g.name, '—') AS group_name, p.amount,
                   date(p.created_at) AS created_date,
                   p.due_date
            FROM payments p
            LEFT JOIN clients c ON c.client_id = p.client_id
            LEFT JOIN groups g ON g.group_id = p.group_id
            WHERE p.status = 'deferred'
              AND p.method = 'defer'
              AND date(p.created_at) BETWEEN ? AND ?
            ORDER BY p.created_at DESC, p.pay_id DESC
            LIMIT 10
            """,
            (date_from, date_to),
        )
        latest = cur.fetchall()

        cur = conn.execute(
            """
            SELECT p.pay_id, COALESCE(c.full_name, '—') AS client_name,
                   COALESCE(g.name, '—') AS group_name, p.amount,
                   date(p.created_at) AS created_date
            FROM payments p
            LEFT JOIN clients c ON c.client_id = p.client_id
            LEFT JOIN groups g ON g.group_id = p.group_id
            WHERE p.status = 'deferred'
              AND p.method = 'defer'
              AND date(p.created_at) <= date(?, ?)
            ORDER BY p.created_at ASC, p.pay_id ASC
            """,
            (today, f"-{overdue_days} days"),
        )
        overdue = cur.fetchall()

    return (
        _int_or_zero(total_count),
        _int_or_zero(total_amount),
        latest,
        overdue,
    )


def list_deferred_payments(
    db_path: str, date_from: str, date_to: str
) -> List[Tuple[str, str, str, int, str, Optional[str]]]:
    with sqlite3.connect(db_path) as conn:
        cur = conn.execute(
            """
            SELECT date(p.created_at) AS created_date,
                   COALESCE(c.full_name, '—') AS client_name,
                   COALESCE(g.name, '—') AS group_name,
                   p.amount,
                   p.purpose,
                   p.due_date
            FROM payments p
            LEFT JOIN clients c ON c.client_id = p.client_id
            LEFT JOIN groups g ON g.group_id = p.group_id
            WHERE p.status = 'deferred'
              AND p.method = 'defer'
              AND date(p.created_at) BETWEEN ? AND ?
            ORDER BY p.created_at DESC, p.pay_id DESC
            """,
            (date_from, date_to),
        )
        return cur.fetchall()


def list_overdue_deferred_payments(
    db_path: str, today: str, overdue_days: int
) -> List[Tuple[str, str, str, int]]:
    with sqlite3.connect(db_path) as conn:
        cur = conn.execute(
            """
            SELECT date(p.created_at) AS created_date,
                   COALESCE(c.full_name, '—') AS client_name,
                   COALESCE(g.name, '—') AS group_name,
                   p.amount
            FROM payments p
            LEFT JOIN clients c ON c.client_id = p.client_id
            LEFT JOIN groups g ON g.group_id = p.group_id
            WHERE p.status = 'deferred'
              AND p.method = 'defer'
              AND date(p.created_at) <= date(?, ?)
            ORDER BY p.created_at ASC, p.pay_id ASC
            """,
            (today, f"-{overdue_days} days"),
        )
        return cur.fetchall()


def build_excel_report(
    db_path: str,
    date_from: str,
    date_to: str,
    today: str,
    overdue_days: int,
) -> bytes:
    revenue_summary = get_revenue_summary(db_path, date_from, date_to)
    expense_summary = get_expense_summary(db_path, date_from, date_to)
    attendance = get_attendance_summary(db_path, date_from, date_to)
    profit = revenue_summary.total - expense_summary.total

    paid_payments = list_paid_payments(db_path, date_from, date_to)
    expenses = list_expenses_for_period(db_path, date_from, date_to)
    deferred = list_deferred_payments(db_path, date_from, date_to)
    overdue = list_overdue_deferred_payments(db_path, today, overdue_days)

    wb = Workbook()
    ws_summary = wb.active
    ws_summary.title = "Summary"

    _append_header(ws_summary, ["Показатель", "Сумма"])
    ws_summary.append(["Выручка", revenue_summary.total])
    ws_summary.append(["Расходы", expense_summary.total])
    ws_summary.append(["Прибыль", profit])
    ws_summary.append([])
    ws_summary.append(["Период", f"{date_from} — {date_to}"])

    ws_revenue = wb.create_sheet("Revenue")
    _append_header(ws_revenue, ["Метод", "Сумма"])
    ws_revenue.append(["Наличные", revenue_summary.cash])
    ws_revenue.append(["Перевод", revenue_summary.transfer])
    ws_revenue.append(["QR", revenue_summary.qr])
    ws_revenue.append([])
    _append_header(ws_revenue, ["Дата", "Клиент", "Группа", "Тип", "Сумма", "Метод"])
    for row in paid_payments:
        ws_revenue.append(list(row))

    ws_expenses = wb.create_sheet("Expenses")
    _append_header(ws_expenses, ["Метод", "Сумма"])
    ws_expenses.append(["Наличные", expense_summary.cash])
    ws_expenses.append(["Перевод", expense_summary.transfer])
    ws_expenses.append(["QR", expense_summary.qr])
    ws_expenses.append([])
    _append_header(ws_expenses, ["Категория", "Сумма"])
    for name, amount in expense_summary.categories:
        ws_expenses.append([name, amount])
    if expense_summary.other_amount:
        ws_expenses.append(["Прочие", expense_summary.other_amount])
    ws_expenses.append([])
    _append_header(ws_expenses, ["Дата", "Категория", "Сумма", "Метод", "Комментарий"])
    for row in expenses:
        ws_expenses.append(list(row))

    ws_attendance = wb.create_sheet("Attendance")
    _append_header(ws_attendance, ["Статус", "Количество"])
    ws_attendance.append(["booked", attendance.booked])
    ws_attendance.append(["attended", attendance.attended])
    ws_attendance.append(["noshow", attendance.noshow])
    ws_attendance.append(["cancelled", attendance.cancelled])
    ws_attendance.append([])
    ws_attendance.append(["Период", f"{date_from} — {date_to}"])

    ws_defers = wb.create_sheet("Defers")
    _append_header(ws_defers, ["Дата", "Клиент", "Группа", "Сумма", "Тип", "Срок оплаты"])
    for row in deferred:
        ws_defers.append(list(row))
    ws_defers.append([])
    _append_header(ws_defers, ["Просрочено с", "Клиент", "Группа", "Сумма"])
    for row in overdue:
        ws_defers.append(list(row))

    for ws in wb.worksheets:
        _auto_fit_columns(ws)

    buffer = BytesIO()
    wb.save(buffer)
    return buffer.getvalue()


def _append_header(ws, values: Iterable[str]) -> None:
    row_values = list(values)
    ws.append(row_values)
    row_idx = ws.max_row
    bold = Font(bold=True)
    for col_idx in range(1, len(row_values) + 1):
        ws.cell(row=row_idx, column=col_idx).font = bold


def _auto_fit_columns(ws) -> None:
    for column_cells in ws.columns:
        length = 0
        col_letter = get_column_letter(column_cells[0].column)
        for cell in column_cells:
            if cell.value is None:
                continue
            value_len = len(str(cell.value))
            if value_len > length:
                length = value_len
        ws.column_dimensions[col_letter].width = min(max(length + 2, 10), 60)
